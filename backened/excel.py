from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_excel(mto_data, ai_data, output_file):
    """
    Creates a professional MTO Excel file.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Material Take-Off"

    # -----------------------------
    # Styles
    # -----------------------------

    title_font = Font(size=18, bold=True)

    heading_font = Font(size=12, bold=True)

    white_font = Font(color="FFFFFF", bold=True)

    blue_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    gray_fill = PatternFill(
        start_color="DDDDDD",
        end_color="DDDDDD",
        fill_type="solid"
    )

    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    thin = Side(style="thin")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    # -----------------------------
    # Title
    # -----------------------------

    ws.merge_cells("A1:D1")

    ws["A1"] = "Material Take-Off (MTO)"

    ws["A1"].font = title_font
    ws["A1"].alignment = center

    # -----------------------------
    # Drawing Information
    # -----------------------------

    ws["A3"] = "Line Number"
    ws["B3"] = ai_data.get("line_number", "")

    ws["A4"] = "Pipe Size"
    ws["B4"] = ai_data.get("pipe_size", "")

    ws["A5"] = "Material"
    ws["B5"] = ai_data.get("material", "")

    ws["A6"] = "Pipe Length"
    ws["B6"] = ai_data.get("pipe_length", "")

    ws["A7"] = "Schedule"
    ws["B7"] = ai_data.get("pipe_schedule", "")

    for cell in ["A3", "A4", "A5", "A6", "A7"]:
        ws[cell].font = heading_font
        ws[cell].fill = gray_fill

    # -----------------------------
    # Table Header
    # -----------------------------

    start_row = 10

    headers = [
        "Item",
        "Description",
        "Size",
        "Quantity"
    ]

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(
            row=start_row,
            column=col
        )

        cell.value = header
        cell.font = white_font
        cell.fill = blue_fill
        cell.alignment = center
        cell.border = border

    # -----------------------------
    # MTO Rows
    # -----------------------------

    row = start_row + 1

    for item in mto_data:

        ws.cell(row=row, column=1).value = item.get("Item", "")
        ws.cell(row=row, column=2).value = item.get("Description", "")
        ws.cell(row=row, column=3).value = item.get("Size", "")
        ws.cell(row=row, column=4).value = item.get("Quantity", "")

        for col in range(1, 5):

            c = ws.cell(row=row, column=col)

            c.border = border
            c.alignment = center

        row += 1

    # -----------------------------
    # Fixed Column Widths
    # -----------------------------

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15

    # -----------------------------
    # Save
    # -----------------------------

    wb.save(output_file)